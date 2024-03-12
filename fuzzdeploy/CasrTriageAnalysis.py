import os
import re

from openpyxl.styles import Font, PatternFill

from . import utility
from .Builder import Builder
from .constants import *
from .CPUAllocator import CPUAllocator
from .ExcelManager import ExcelManager


class CasrTriageAnalysis:
    fuzzer_colors = {
        "aflplusplus": "00A0B0",
        "htfuzz": "FBD26A",
    }
    display_fields = [
        FUZZER,
        REPEAT,
        "unique_line",
        "casr_dedup",
        "casr_dedup_cluster",
    ]
    Severity = {
        "EXPLOITABLE": (
            "SegFaultOnPc",
            "ReturnAv",
            "BranchAv",
            "CallAv",
            "DestAv",
            "BranchAvTainted",
            "CallAvTainted",
            "DestAvTainted",
            "heap-buffer-overflow(write)",
            "global-buffer-overflow(write)",
            "stack-use-after-scope(write)",
            "stack-use-after-return(write)",
            "stack-buffer-overflow(write)",
            "stack-buffer-underflow(write)",
            "heap-use-after-free(write)",
            "container-overflow(write)",
            "param-overlap",
        ),
        "PROBABLY_EXPLOITABLE": (
            "BadInstruction",
            "SegFaultOnPcNearNull",
            "BranchAvNearNull",
            "CallAvNearNull",
            "HeapError",
            "StackGuard",
            "DestAvNearNull",
            "heap-buffer-overflow",
            "global-buffer-overflow",
            "stack-use-after-scope",
            "use-after-poison",
            "stack-use-after-return",
            "stack-buffer-overflow",
            "stack-buffer-underflow",
            "heap-use-after-free",
            "container-overflow",
            "negative-size-param",
            "calloc-overflow",
            "readllocarray-overflow",
            "pvalloc-overflow",
            "overwrites-const-input",
        ),
        "NOT_EXPLOITABLE": (
            "SourceAv",
            "AbortSignal",
            "AccessViolation",
            "SourceAvNearNull",
            "SafeFunctionCheck",
            "FPE",
            "StackOverflow",
            "double-free",
            "bad-free",
            "alloc-dealloc-mismatch",
            "heap-buffer-overflow(read)",
            "global-buffer-overflow(read)",
            "stack-use-after-scope(read)",
            "stack-use-after-return(read)",
            "stack-buffer-overflow(read)",
            "stack-buffer-underflow(read)",
            "heap-use-after-free(read)",
            "container-overflow(read)",
            "initialization-order-fiasco",
            "new-delete-type-mismatch",
            "bad-malloc_usable_size",
            "odr-violation",
            "memory-leaks",
            "invalid-allocation-alignment",
            "invalid-aligned-alloc-alignment",
            "invalid-posix-memalign-alignment",
            "allocation-size-too-big",
            "out-of-memory",
            "fuzz target exited",
            "timeout",
        ),
    }

    @staticmethod
    def is_heap_related_bug(bug_type):
        if "stack" in bug_type.lower():
            return False
        return bug_type not in (
            "param-overlap",
            "BadInstruction",
            "overwrites-const-input",
            "AbortSignal",
            "SafeFunctionCheck",
            "FPE",
            "initialization-order-fiasco",
            "odr-violation",
            "fuzz target exited",
            "timeout",
        )

    @staticmethod
    @utility.time_count("CRASH TRIAGE BY CASR@https://github.com/ispras/casr DONE!")
    def obtain(WORK_DIR):
        assert os.path.exists(WORK_DIR), f"{WORK_DIR} not exists"
        WORK_DIR_TRIAGE_BY_CASR = os.path.join(WORK_DIR, TRIAGE_BY_CASR)
        # check if images exist
        TARGETS = set()
        for fuzzer, target, repeat, test_path in utility.get_workdir_paths_by(WORK_DIR):
            TARGETS.add(target)
        Builder.build_imgs(FUZZERS=["casr"], TARGETS=list(TARGETS))

        # calculate crashes sum
        def get_triaged_crashes_num(triaged_path):
            triaged_num = 0
            if os.path.exists(os.path.join(triaged_path, "failed")):
                triaged_num += len(os.listdir(os.path.join(triaged_path, "failed")))
            if os.path.exists(os.path.join(triaged_path, "reports")):
                triaged_num += len(os.listdir(os.path.join(triaged_path, "reports")))
            return triaged_num

        crashes_sum = {}
        current_crashes_num = {}
        untriaged_paths = []
        for fuzzer, target, repeat, test_path in utility.get_workdir_paths_by(WORK_DIR):
            assert os.path.exists(
                os.path.join(test_path, TARGET_ARGS)
            ), f"{TARGET_ARGS} not found in {test_path}"
            fuzzer_stats_path = utility.search_item(test_path, "FILE", FUZZER_STATS)
            if fuzzer_stats_path is None:
                utility.console.print(
                    f"[yellow]Warning: {FUZZER_STATS} not found in {test_path}, maybe fine.[/yellow]"
                )
            triage_by_casr = os.path.join(
                WORK_DIR_TRIAGE_BY_CASR, fuzzer, target, repeat
            )
            triaged_num = get_triaged_crashes_num(triage_by_casr)
            for foldername, subfolders, filenames in os.walk(test_path):
                if "crashes" in subfolders:
                    files = os.listdir(os.path.join(foldername, "crashes"))
                    if "README.txt" in files:
                        files.remove("README.txt")
                    crashes_sum[f"{fuzzer}/{target}/{repeat}"] = len(files)
                    break
            if f"{fuzzer}/{target}/{repeat}" not in crashes_sum:
                continue
            if triaged_num != crashes_sum[f"{fuzzer}/{target}/{repeat}"]:
                untriaged_paths.append(test_path)
            elif triaged_num == 0 and (
                not os.path.exists(
                    os.path.join(triage_by_casr, "summary_by_unique_line")
                )
            ):
                untriaged_paths.append(test_path)

        with utility.Progress(
            utility.SpinnerColumn(spinner_name="arrow3"),
            utility.TextColumn("[progress.description]{task.description}"),
            utility.BarColumn(),
            utility.TextColumn("[bold]{task.completed} / {task.total}"),
            utility.TimeElapsedColumn(),
            transient=True,
        ) as progress:
            cpu_allocator = CPUAllocator()
            last_triaged_crashes_num = 0

            def update_progress(progress, last_triaged_crashes_num):
                for test_path in untriaged_paths:
                    fuzzer, target, repeat = utility.parse_path_by(test_path)
                    triage_by_casr = os.path.join(
                        WORK_DIR_TRIAGE_BY_CASR, fuzzer, target, repeat
                    )
                    if crashes_sum[
                        f"{fuzzer}/{target}/{repeat}"
                    ] == current_crashes_num.get(f"{fuzzer}/{target}/{repeat}", 0):
                        continue
                    current_crashes_num[
                        f"{fuzzer}/{target}/{repeat}"
                    ] = get_triaged_crashes_num(triage_by_casr)
                triaged_crashes_num = sum(current_crashes_num.values())
                progress.update(
                    triage_task,
                    advance=triaged_crashes_num - last_triaged_crashes_num,
                )
                return triaged_crashes_num

            triage_task = progress.add_task(
                "[bold green]Triaging", total=sum(crashes_sum.values())
            )
            for test_path in untriaged_paths:
                fuzzer, target, repeat = utility.parse_path_by(test_path)
                triage_by_casr = os.path.join(
                    WORK_DIR_TRIAGE_BY_CASR, fuzzer, target, repeat
                )
                os.makedirs(triage_by_casr, exist_ok=True)
                while True:
                    cpu_id = cpu_allocator.get_free_cpu(sleep_time=1, time_out=10)
                    last_triaged_crashes_num = update_progress(
                        progress, last_triaged_crashes_num
                    )
                    if cpu_id is not None:
                        break
                container_id = utility.get_cmd_res(
                    f"""
            docker run \
            -itd \
            --rm \
            --volume={test_path}:/shared \
            --volume={triage_by_casr}:/triage_by_casr \
            --cap-add=SYS_PTRACE \
            --security-opt seccomp=unconfined \
            --cpuset-cpus="{cpu_id}" \
            --network=none \
            "casr/{target}" \
            -c '${{SRC}}/triage_by_casr.sh'
                    """
                ).strip()
                cpu_allocator.append(container_id, cpu_id)
            while len(cpu_allocator.container_id_dict) > 0:
                last_triaged_crashes_num = update_progress(
                    progress, last_triaged_crashes_num
                )
                cpu_id = cpu_allocator.get_free_cpu(sleep_time=5, time_out=10)
                if cpu_id is None:
                    continue
                container_id_dict = cpu_allocator.container_id_dict
                if len(container_id_dict) == 0:
                    break
                min_container_id = min(
                    container_id_dict, key=lambda k: len(container_id_dict[k])
                )
                allocated_cpu_ls = cpu_allocator.append(min_container_id, cpu_id)
                utility.get_cmd_res(
                    f"docker update --cpuset-cpus {','.join(allocated_cpu_ls)} {min_container_id} 2>/dev/null"
                )
        triage_results = {}
        for fuzzer, target, repeat, test_path in utility.get_workdir_paths_by(WORK_DIR):
            triage_by_casr = os.path.join(
                WORK_DIR_TRIAGE_BY_CASR, fuzzer, target, repeat
            )
            triage = {
                FUZZER: fuzzer,
                REPEAT: repeat,
                "unique_line": 0,
                "casr_dedup": 0,
                "casr_dedup_cluster": 0,
            }
            if os.path.exists(os.path.join(triage_by_casr, "reports_dedup")):
                triage["casr_dedup"] = len(
                    os.listdir(os.path.join(triage_by_casr, "reports_dedup"))
                )
            if os.path.exists(os.path.join(triage_by_casr, "reports_dedup_cluster")):
                triage["casr_dedup_cluster"] = len(
                    os.listdir(os.path.join(triage_by_casr, "reports_dedup_cluster"))
                )
            summary_by_unique_line_path = os.path.join(
                triage_by_casr, "summary_by_unique_line"
            )
            if os.path.isfile(summary_by_unique_line_path):
                with open(summary_by_unique_line_path) as f:
                    summary_content = f.read()
                data_str = summary_content.split("->")[-1]
                pattern = r"(.+?): (\d+)"
                for match in re.findall(pattern, data_str):
                    key, value = match
                    key = key.strip()
                    value = int(value.strip())
                    triage[key] = value
                    triage["unique_line"] += value
            triage_results.setdefault(target, []).append(triage)
        for target in triage_results.keys():
            triage_results[target] = sorted(
                triage_results[target],
                key=lambda x: (
                    x["unique_line"],
                    x["casr_dedup_cluster"],
                    x["casr_dedup"],
                ),
                reverse=True,
            )
        utility.console.print(f"The results can be found in {WORK_DIR_TRIAGE_BY_CASR}")
        return triage_results

    @staticmethod
    def sort_by_severity(field):
        if field in CasrTriageAnalysis.Severity["EXPLOITABLE"]:
            return 0, field
        elif field in CasrTriageAnalysis.Severity["PROBABLY_EXPLOITABLE"]:
            return 1, field
        return 2, field

    @staticmethod
    def sort_by_severity_and_crashline(bug_field):
        field, crashline = bug_field.split("/", 1)
        return *CasrTriageAnalysis.sort_by_severity(field), crashline

    @staticmethod
    def get_severity_color(field):
        color = "000000"
        if field in CasrTriageAnalysis.Severity["EXPLOITABLE"]:
            color = "800000"
        elif field in CasrTriageAnalysis.Severity["PROBABLY_EXPLOITABLE"]:
            color = "1F497D"
        return color

    @staticmethod
    # @utility.time_count("SAVE TRIAGE BY CASR@https://github.com/ispras/casr DONE!")
    def save(WORK_DIR, OUTPUT_FILE=None, HEAP_RELATED_BUGS_FIELD=False):
        if OUTPUT_FILE is None:
            OUTPUT_FILE = os.path.join(
                WORK_DIR,
                f"{os.path.basename(WORK_DIR)}_triage_by_casr.xlsx",
            )
        triage_results = CasrTriageAnalysis.obtain(WORK_DIR)
        excel_manager = ExcelManager()
        for target in sorted(triage_results.keys()):
            table_data = triage_results[target]
            bug_fields = list(set([_ for item in table_data for _ in item.keys()]))
            bug_fields = [
                item
                for item in bug_fields
                if item not in CasrTriageAnalysis.display_fields
            ]
            bug_fields = sorted(bug_fields, key=CasrTriageAnalysis.sort_by_severity)
            if HEAP_RELATED_BUGS_FIELD:
                display_fields = (
                    CasrTriageAnalysis.display_fields + [HEAP_RELATED_BUGS] + bug_fields
                )
            else:
                display_fields = CasrTriageAnalysis.display_fields + bug_fields
            excel_manager.create_sheet(target)
            # the header of table
            excel_manager.set_sheet_header(
                display_fields,
                [
                    {
                        "Font": Font(
                            bold=True,
                            name="Calibri",
                            size=17,
                            color=CasrTriageAnalysis.get_severity_color(display_field),
                        )
                    }
                    for display_field in display_fields
                ],
            )
            # the rows of table
            for item in table_data:
                if HEAP_RELATED_BUGS_FIELD:
                    item[HEAP_RELATED_BUGS] = 0
                    for bug_field in bug_fields:
                        if (
                            bug_field in item.keys()
                            and CasrTriageAnalysis.is_heap_related_bug(bug_field)
                        ):
                            item[HEAP_RELATED_BUGS] += item[bug_field]
                excel_manager.set_sheet_data(
                    [
                        item[display_field] if display_field in item.keys() else ""
                        for display_field in display_fields
                    ],
                    [
                        {
                            "Fill": PatternFill(
                                fgColor=CasrTriageAnalysis.fuzzer_colors[item[FUZZER]],
                                fill_type="solid",
                            )
                        }
                        if item[FUZZER] in CasrTriageAnalysis.fuzzer_colors.keys()
                        else {}
                        for _ in display_fields
                    ],
                )
        excel_manager.save_workbook(OUTPUT_FILE)
        utility.console.print(
            f"triage by casr@https://github.com/ispras/casr are saved in {OUTPUT_FILE}"
        )
        return OUTPUT_FILE
